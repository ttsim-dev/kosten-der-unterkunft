"""Fetch Gemeinde population and area and write `data/gemeinde_population.arrow`.

Downloads two editions of the Destatis GV-ISys Gemeindeverzeichnis
Jahresausgabe into `bld/` (gitignored), reconciles them against the AGS of
`data/gemeinden.geo.json`, and writes the committed population table.

Re-run to regenerate:

    pixi run python scripts/fetch_gemeinde_population.py

Why two editions. No published Gebietsstand reproduces the 10,980 AGS of the
boundary export exactly, because that export mixes vintages. The 31.12.2023
edition comes closest: it is short exactly three of them and carries four
records — three Hoheits- or Küstengewässergebiete and one gemeindefreies
Gebiet — that the boundary file does not draw. The three missing AGS are
restored from the 31.12.2022 edition through the explicit
`MERGER_REVERSALS` below, and the extra records are dropped by the join
against the boundary AGS.
"""

import json
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from _gemeindeverzeichnis import (
    MergerReversal,
    build_gemeinde_population,
    parse_gv_rows,
)

from kdu.config import BLD, DATA, LEGAL_VINTAGE

SOURCE_URL_TEMPLATE = (
    "https://www.destatis.de/DE/Themen/Laender-Regionen/Regionales/"
    "Gemeindeverzeichnis/Administrativ/Archiv/GVAuszugJ/"
    "{stamp}_Auszug_GV.xlsx?__blob=publicationFile"
)
BASE_REFERENCE_DATE = date(2023, 12, 31)
BACKFILL_REFERENCE_DATE = date(2022, 12, 31)

# Corrections turning the 31.12.2023 edition into the boundary Gebietsstand.
MERGER_REVERSALS: tuple[MergerReversal, ...] = (
    MergerReversal(
        absorbed_ags=("01059101", "01059141"),
        successor_ags="01059126",
        note=(
            "Tastrup and Maasbüll merged into Hürup on 2023-01-01; the boundary "
            "file still draws them, so their 31.12.2022 figures are restored and "
            "netted out of Hürup."
        ),
    ),
    MergerReversal(
        absorbed_ags=("09374451",),
        successor_ags=None,
        note=(
            "Heinersreuther Forst, a gemeindefreies Gebiet with no inhabitants, "
            "was dissolved during 2023 and its area distributed over several "
            "neighbours. Population is zero, so restoring its 31.12.2022 record "
            "double-counts 5.88 km² of area and no inhabitants."
        ),
    ),
)


def main() -> None:
    """Download both editions, reconcile them, and write the committed table."""
    BLD.mkdir(exist_ok=True)
    base = parse_gv_rows(_read_rows(_download(BASE_REFERENCE_DATE)))
    backfill = parse_gv_rows(_read_rows(_download(BACKFILL_REFERENCE_DATE)))
    boundary_ags = _boundary_ags(DATA / "gemeinden.geo.json")

    population = build_gemeinde_population(
        base=base,
        backfill=backfill,
        boundary_ags=boundary_ags,
        reversals=MERGER_REVERSALS,
        base_reference_date=BASE_REFERENCE_DATE,
        backfill_reference_date=BACKFILL_REFERENCE_DATE,
    )

    out_path = DATA / "gemeinde_population.arrow"
    population.to_feather(out_path)
    print(
        f"{len(population)} Gemeinden, Gebietsstand "
        f"{LEGAL_VINTAGE.gebietsstand.isoformat()}, "
        f"{population['population'].sum():,} inhabitants → {out_path}"
    )
    for reversal in MERGER_REVERSALS:
        print(f"  reversal {', '.join(reversal.absorbed_ags)}: {reversal.note}")


def _download(reference_date: date) -> Path:
    """Fetch one Jahresausgabe into `bld/`, reusing an earlier download."""
    stamp = reference_date.strftime("%d%m%Y")
    path = BLD / f"gemeindeverzeichnis_{stamp}.xlsx"
    if not path.exists():
        url = SOURCE_URL_TEMPLATE.format(stamp=stamp)
        print(f"downloading {url}")
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "kdu-research/1.0"},
        )
        with urllib.request.urlopen(request) as response:
            path.write_bytes(response.read())
    return path


def _read_rows(path: Path) -> list[tuple[Any, ...]]:
    """Return every cell tuple of the workbook's Onlineprodukt sheet."""
    workbook = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[-1]]
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _boundary_ags(path: Path) -> pd.Series:
    """Derive the eight-digit AGS of every geometry from its 12-digit key."""
    raw = json.loads(path.read_text(encoding="utf-8"))
    codes = pd.Series(
        [feature["properties"]["gem_code"] for feature in raw["features"]],
        dtype="string",
    )
    return codes.str[:5] + codes.str[-3:]


if __name__ == "__main__":
    main()
